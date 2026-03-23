"""
Reads Dataset.xlsx, loads each HuggingFace dataset listed in every sheet,
randomly samples 10 question-answer pairs, and writes a unified CSV file
with columns: dataset_name, domain, question, answer.

The CSV is updated after every dataset so progress is never lost if the
script is interrupted.

Usage:
    pip install datasets openpyxl pandas
    python sample_datasets.py
"""

import random
import pandas as pd
from openpyxl import load_workbook
from datasets import load_dataset

from huggingface_hub import login
login(token="") # Comment this if you don't have a token

EXCEL_PATH = "/content/Dataset.xlsx"   # path to your Excel file
OUTPUT_CSV = "/content/Knowledge_Sampled_QA.csv"
N_SAMPLES  = 10
SEED       = 42

random.seed(SEED)


def load_sheet_entries(excel_path: str) -> list[dict]:
    """Parse all sheets and return a list of dataset entry dicts."""
    wb = load_workbook(excel_path, read_only=True)
    entries = []
    for domain in wb.sheetnames:
        ws = wb[domain]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip() if c else "" for c in rows[0]]
        col = {name: idx for idx, name in enumerate(header)}
        for row in rows[1:]:
            name     = row[col["name"]]     if "name"     in col else None
            provider = row[col["provider"]] if "provider" in col else None
            q_col    = row[col["inputs"]]   if "inputs"   in col else None
            a_col    = row[col["outputs"]]  if "outputs"  in col else None
            load_arg = row[col["load"]]     if "load"     in col else None
            if not name or str(provider).strip().upper() != "HF":
                continue
            entries.append({
                "domain":       domain,
                "dataset_name": str(name).strip(),
                "q_col":        str(q_col).strip()    if q_col    else None,
                "a_col":        str(a_col).strip()    if a_col    else None,
                "load_arg":     str(load_arg).strip() if load_arg and load_arg != "None" else None,
            })
    return entries


def sample_from_dataset(entry: dict, n: int = N_SAMPLES) -> list[dict]:
    """Load a HuggingFace dataset and return n random Q-A pairs."""
    name     = entry["dataset_name"]
    q_col    = entry["q_col"]
    a_col    = entry["a_col"]
    load_arg = entry["load_arg"]
    domain   = entry["domain"]

    print(f"  Loading: {name}  (config={load_arg}, Q='{q_col}', A='{a_col}')")

    try:
        def _load(trust: bool):
            kwargs = dict(trust_remote_code=trust) if trust else {}
            if load_arg:
                return load_dataset(name, load_arg, **kwargs)
            return load_dataset(name, **kwargs)

        try:
            ds = _load(trust=False)
        except Exception:
            ds = _load(trust=True)

        # Pick a split (prefer train)
        if hasattr(ds, "keys"):
            split_name = "train" if "train" in ds else list(ds.keys())[0]
            split = ds[split_name]
        else:
            split = ds

        total = len(split)
        if total == 0:
            print("    [SKIP] Dataset is empty.")
            return []

        available = split.column_names
        if q_col not in available or a_col not in available:
            print(f"    [SKIP] Columns not found. Available: {available}")
            return []

        indices = random.sample(range(total), min(n, total))
        rows = split.select(indices)

        samples = []
        for row in rows:
            q = str(row[q_col]).strip()
            a = str(row[a_col]).strip()
            if q and a:
                samples.append({
                    "dataset_name": name,
                    "domain":       domain,
                    "question":     q,
                    "answer":       a,
                })
        print(f"    [OK] Sampled {len(samples)} rows.")
        return samples

    except Exception as exc:
        print(f"    [ERROR] {exc}")
        return []


def append_to_csv(rows: list[dict], output_path: str, write_header: bool) -> None:
    """Append a batch of rows to the CSV; write header only on the first call."""
    if not rows:
        return
    df = pd.DataFrame(rows, columns=["dataset_name", "domain", "question", "answer"])
    df.to_csv(output_path, mode="a", index=False, encoding="utf-8", header=write_header)


def main():
    print(f"Reading {EXCEL_PATH} ...")
    entries = load_sheet_entries(EXCEL_PATH)
    print(f"Found {len(entries)} dataset entries across all domains.\n")

    total_rows = 0
    first_write = True

    for entry in entries:
        print(f"[{entry['domain'].upper()}] {entry['dataset_name']}")
        rows = sample_from_dataset(entry)
        if rows:
            append_to_csv(rows, OUTPUT_CSV, write_header=first_write)
            first_write = False
            total_rows += len(rows)
            print(f"    [SAVED] {OUTPUT_CSV} -- {total_rows} rows total so far.")
        print()

    if total_rows == 0:
        print("No data collected -- check errors above.")
    else:
        print(f"\nDone! {total_rows} rows written to '{OUTPUT_CSV}'.")


if __name__ == "__main__":
    main()
